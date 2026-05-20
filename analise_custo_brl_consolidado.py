#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================================
# 1. PARSE BINANCE CSV - CONVERSÕES BRL
# ============================================================================
binance_file = "Binance-Histórico-de-Transações-202605111552(UTC--3).csv"
binance_brl_conversions = []
binance_by_token = defaultdict(lambda: {'qty': 0, 'brl_spent': 0, 'conversions': []})

with open(binance_file, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
    for i in range(1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(',')]
        if len(parts) < 5:
            continue

        tempo = parts[1]
        operacao = parts[3]
        moeda = parts[4]
        try:
            valor = float(parts[5].replace(',', '.'))
        except:
            continue

        # Binance Convert: BRL negativo (saída), token positivo (entrada)
        if operacao == 'Binance Convert' and moeda != 'BRL' and valor > 0:
            # Procura linha anterior com BRL negativo
            if i > 1:
                prev_line = lines[i-1].strip().split(',')
                if len(prev_line) >= 6:
                    prev_moeda = prev_line[4].strip()
                    try:
                        prev_valor = float(prev_line[5].replace(',', '.'))
                    except:
                        continue

                    if prev_moeda == 'BRL' and prev_valor < 0:
                        brl_spent = abs(prev_valor)
                        binance_brl_conversions.append({
                            'date': tempo,
                            'operation': 'Binance Convert',
                            'token': moeda,
                            'qty': valor,
                            'brl_spent': brl_spent,
                            'rate': brl_spent / valor if valor > 0 else 0
                        })
                        binance_by_token[moeda]['qty'] += valor
                        binance_by_token[moeda]['brl_spent'] += brl_spent
                        binance_by_token[moeda]['conversions'].append({
                            'date': tempo,
                            'qty': valor,
                            'brl': brl_spent,
                            'rate': brl_spent / valor if valor > 0 else 0
                        })

        # Transaction Buy: moeda é a criptomoeda, procura linha com Transaction Spend BRL
        if operacao == 'Transaction Buy' and moeda != 'BRL' and valor > 0:
            # Procura Transaction Spend em BRL na mesma série de linhas (mesmo timestamp)
            for j in range(max(0, i-3), min(len(lines), i+3)):
                check_line = lines[j].strip().split(',')
                if len(check_line) >= 6 and j != i:
                    check_op = check_line[3].strip()
                    check_moeda = check_line[4].strip()
                    try:
                        check_valor = float(check_line[5].replace(',', '.'))
                    except:
                        continue

                    if check_op == 'Transaction Spend' and check_moeda == 'BRL' and check_valor < 0:
                        brl_spent = abs(check_valor)
                        # Verifica se não está duplicado
                        already_recorded = any(
                            c['date'] == tempo and c['token'] == moeda and abs(c['brl_spent'] - brl_spent) < 0.01
                            for c in binance_brl_conversions
                        )
                        if not already_recorded:
                            binance_brl_conversions.append({
                                'date': tempo,
                                'operation': 'Transaction Buy',
                                'token': moeda,
                                'qty': valor,
                                'brl_spent': brl_spent,
                                'rate': brl_spent / valor if valor > 0 else 0
                            })
                            binance_by_token[moeda]['qty'] += valor
                            binance_by_token[moeda]['brl_spent'] += brl_spent
                            binance_by_token[moeda]['conversions'].append({
                                'date': tempo,
                                'qty': valor,
                                'brl': brl_spent,
                                'rate': brl_spent / valor if valor > 0 else 0
                            })
                        break

# ============================================================================
# 2. PARSE OKX CSV - CONVERSÕES BRL
# ============================================================================
okx_file = "Mjk2NTE5NTE=~2023-05-09~2026-05-09~UTC+8~ea1e6f1509d9e505a9af7c3a52c37475_UnifiedBillHistory/OKX Trading History_2023-05-09~2026-05-09~UTC+8~ea1e6f1509d9e505a9af7c3a52c37475.csv"
okx_brl_conversions = []
okx_by_token = defaultdict(lambda: {'qty': 0, 'brl_spent': 0, 'conversions': []})

with open(okx_file, 'r', encoding='utf-8-sig') as f:
    next(f)  # Skip the BOM header
    reader = csv.DictReader(f)
    for row in reader:
        trade_type = row.get('Trade Type', '').strip()
        symbol = row.get('Symbol', '').strip()
        action = row.get('Action', '').strip()

        if trade_type == 'Spot' and symbol == 'USDT-BRL' and action == 'Buy':
            try:
                amount = float(row.get('Amount', '0'))
                price = float(row.get('Filled Price', '0'))
                brl_spent = amount * price
                tempo = row.get('Time', '')

                okx_brl_conversions.append({
                    'date': tempo,
                    'operation': 'USDT-BRL Buy',
                    'token': 'USDT',
                    'qty': amount,
                    'brl_spent': brl_spent,
                    'rate': price
                })
                okx_by_token['USDT']['qty'] += amount
                okx_by_token['USDT']['brl_spent'] += brl_spent
                okx_by_token['USDT']['conversions'].append({
                    'date': tempo,
                    'qty': amount,
                    'brl': brl_spent,
                    'rate': price
                })
            except:
                pass

# ============================================================================
# 3. CONSOLIDAR E CALCULAR
# ============================================================================
all_conversions = binance_brl_conversions + okx_brl_conversions
all_conversions.sort(key=lambda x: datetime.strptime(x['date'].split()[0], '%y-%m-%d') if 'y-' in x['date'] else datetime.now())

# Consolidado por token
consolidated = {}
for token in set(list(binance_by_token.keys()) + list(okx_by_token.keys())):
    b = binance_by_token[token]
    o = okx_by_token[token]
    total_qty = b['qty'] + o['qty']
    total_brl = b['brl_spent'] + o['brl_spent']

    consolidated[token] = {
        'binance_qty': b['qty'],
        'binance_brl': b['brl_spent'],
        'okx_qty': o['qty'],
        'okx_brl': o['brl_spent'],
        'total_qty': total_qty,
        'total_brl': total_brl,
        'avg_rate': total_brl / total_qty if total_qty > 0 else 0,
        'binance_conversions': b['conversions'],
        'okx_conversions': o['conversions']
    }

# ============================================================================
# 4. GERAR EXCEL
# ============================================================================
wb = Workbook()
ws = wb.active
ws.title = "Resumo"

# Cabeçalho
ws['A1'] = "CONSOLIDAÇÃO DE CUSTO DE AQUISIÇÃO - BRL"
ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws.merge_cells('A1:H1')

ws['A2'] = "Relatório gerado em: " + datetime.now().strftime('%d/%m/%Y %H:%M')
ws['A2'].font = Font(name='Arial', size=10, italic=True)
ws.merge_cells('A2:H2')

# Resumo por token
row = 4
ws[f'A{row}'] = "RESUMO POR TOKEN"
ws[f'A{row}'].font = Font(size=12, bold=True)
ws[f'A{row}'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
ws.merge_cells(f'A{row}:H{row}')

row = 5
headers = ['Token', 'Binance Qty', 'Binance BRL', 'OKX Qty', 'OKX BRL', 'Total Qty', 'Total BRL', 'Taxa Média (BRL/unit)']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=row, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

row = 6
for token in sorted(consolidated.keys()):
    data = consolidated[token]
    ws[f'A{row}'] = token
    ws[f'B{row}'] = data['binance_qty']
    ws[f'B{row}'].number_format = '0.00000000'
    ws[f'C{row}'] = data['binance_brl']
    ws[f'C{row}'].number_format = '#,##0.00'
    ws[f'D{row}'] = data['okx_qty']
    ws[f'D{row}'].number_format = '0.00000000'
    ws[f'E{row}'] = data['okx_brl']
    ws[f'E{row}'].number_format = '#,##0.00'
    ws[f'F{row}'] = data['total_qty']
    ws[f'F{row}'].number_format = '0.00000000'
    ws[f'G{row}'] = data['total_brl']
    ws[f'G{row}'].number_format = '#,##0.00'
    ws[f'H{row}'] = data['avg_rate']
    ws[f'H{row}'].number_format = '0.00000'
    row += 1

# Totais
ws[f'A{row}'] = "TOTAL"
ws[f'A{row}'].font = Font(bold=True)
total_qty = sum(d['total_qty'] for d in consolidated.values())
total_brl = sum(d['total_brl'] for d in consolidated.values())
ws[f'F{row}'] = total_qty
ws[f'F{row}'].font = Font(bold=True)
ws[f'F{row}'].number_format = '0.00000000'
ws[f'G{row}'] = total_brl
ws[f'G{row}'].font = Font(bold=True)
ws[f'G{row}'].number_format = '#,##0.00'

# Ajustar larguras
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 14
ws.column_dimensions['H'].width = 18

# ============================================================================
# 5. FOLHA 2: HISTÓRICO BINANCE
# ============================================================================
ws2 = wb.create_sheet("Binance - Histórico")
ws2['A1'] = "BINANCE - HISTÓRICO DE CONVERSÕES BRL"
ws2['A1'].font = Font(size=12, bold=True, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws2.merge_cells('A1:E1')

ws2['A3'] = "Data"
ws2['B3'] = "Token"
ws2['C3'] = "Quantidade"
ws2['D3'] = "BRL Gasto"
ws2['E3'] = "Taxa (BRL/unit)"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws2[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws2[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

row = 4
for conv in sorted([c for c in binance_brl_conversions], key=lambda x: x['date']):
    ws2[f'A{row}'] = conv['date']
    ws2[f'B{row}'] = conv['token']
    ws2[f'C{row}'] = conv['qty']
    ws2[f'C{row}'].number_format = '0.00000000'
    ws2[f'D{row}'] = conv['brl_spent']
    ws2[f'D{row}'].number_format = '#,##0.00'
    ws2[f'E{row}'] = conv['rate']
    ws2[f'E{row}'].number_format = '0.00000'
    row += 1

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 16

# ============================================================================
# 6. FOLHA 3: HISTÓRICO OKX
# ============================================================================
ws3 = wb.create_sheet("OKX - Histórico")
ws3['A1'] = "OKX - HISTÓRICO DE CONVERSÕES USDT-BRL"
ws3['A1'].font = Font(size=12, bold=True, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws3.merge_cells('A1:E1')

ws3['A3'] = "Data"
ws3['B3'] = "Token"
ws3['C3'] = "Quantidade"
ws3['D3'] = "BRL Gasto"
ws3['E3'] = "Taxa (BRL/unit)"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws3[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws3[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

row = 4
for conv in sorted([c for c in okx_brl_conversions], key=lambda x: x['date']):
    ws3[f'A{row}'] = conv['date']
    ws3[f'B{row}'] = conv['token']
    ws3[f'C{row}'] = conv['qty']
    ws3[f'C{row}'].number_format = '0.00000000'
    ws3[f'D{row}'] = conv['brl_spent']
    ws3[f'D{row}'].number_format = '#,##0.00'
    ws3[f'E{row}'] = conv['rate']
    ws3[f'E{row}'].number_format = '0.00000'
    row += 1

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16

# ============================================================================
# 7. FOLHA 4: EVOLUÇÃO BRL/USDT
# ============================================================================
ws4 = wb.create_sheet("Evolução BRL-USDT")
ws4['A1'] = "EVOLUÇÃO DA TAXA BRL/USDT NO TEMPO"
ws4['A1'].font = Font(size=12, bold=True, color='FFFFFF')
ws4['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws4.merge_cells('A1:C1')

ws4['A3'] = "Data"
ws4['B3'] = "Taxa BRL/USDT"
ws4['C3'] = "Quantidade (USDT)"

for col in ['A', 'B', 'C']:
    ws4[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws4[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

row = 4
usdt_conversions = sorted(
    okx_by_token['USDT']['conversions'] if 'USDT' in okx_by_token else [],
    key=lambda x: datetime.strptime(x['date'].split()[0], '%Y-%m-%d') if '-' in x['date'] else datetime.now()
)
for conv in usdt_conversions:
    ws4[f'A{row}'] = conv['date']
    ws4[f'B{row}'] = conv['rate']
    ws4[f'B{row}'].number_format = '0.00000'
    ws4[f'C{row}'] = conv['qty']
    ws4[f'C{row}'].number_format = '0.00000000'
    row += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 18

# Salvar
wb.save('Custo_BRL_Consolidado_Lucas.xlsx')
print("[OK] COMPLETO: Custo_BRL_Consolidado_Lucas.xlsx")
print(f"\nRESUMO:")
print(f"  Tokens únicos: {len(consolidated)}")
print(f"  Conversões Binance: {len(binance_brl_conversions)}")
print(f"  Conversões OKX: {len(okx_brl_conversions)}")
print(f"  Total BRL investido: R$ {total_brl:,.2f}")
print(f"\nPOR TOKEN:")
for token in sorted(consolidated.keys()):
    d = consolidated[token]
    print(f"  {token}: {d['total_qty']:.8f} @ R$ {d['avg_rate']:.5f} = R$ {d['total_brl']:,.2f}")

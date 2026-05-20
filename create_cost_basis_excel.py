#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "Resumo"

# FOLHA 1: RESUMO EXECUTIVO
ws['A1'] = "BAROLO CAPITAL - CUSTO DE AQUISICAO EM BRL"
ws['A1'].font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws.merge_cells('A1:D1')

ws['A2'] = "Relatorio gerado em: " + datetime.now().strftime('%d/%m/%Y %H:%M')
ws['A2'].font = Font(name='Arial', size=10, italic=True)
ws.merge_cells('A2:D2')

# Resumo
ws['A4'] = "RESUMO CONSOLIDADO"
ws['A4'].font = Font(name='Arial', size=12, bold=True)
ws['A4'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
ws.merge_cells('A4:D4')

ws['A6'] = "Metrica"
ws['B6'] = "Valor (BRL)"
for col in ['A', 'B']:
    ws[f'{col}6'].font = Font(bold=True, color='FFFFFF')
    ws[f'{col}6'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

data_summary = [
    ("BRL Total Transferido", 7990.00),
    ("BRL Total Investido em Cripto", 5815.50),
    ("BRL Saldo Disponivel", 2174.50),
    ("", ""),
    ("Fonte OKX (USDT)", 5065.50),
    ("Fonte Binance (USDC)", 750.00),
]

row = 7
for label, value in data_summary:
    if label == "":
        row += 1
    else:
        ws[f'A{row}'] = label
        if value != "":
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

# FOLHA 1: CUSTO POR TOKEN
ws['A15'] = "CUSTO DE AQUISICAO POR TOKEN"
ws['A15'].font = Font(name='Arial', size=12, bold=True)
ws['A15'].fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
ws.merge_cells('A15:E15')

ws['A17'] = "Token"
ws['B17'] = "Total BRL"
ws['C17'] = "Quantidade"
ws['D17'] = "Preco/Unidade (BRL)"
ws['E17'] = "# Transacoes"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws[f'{col}17'].font = Font(bold=True, color='FFFFFF')
    ws[f'{col}17'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

tokens_data = [
    ("USDT", 5065.50, 989.52188000, 5.11914, 14),
    ("USDC", 750.00, 134.35758172, 5.58212, 2),
]

row = 18
for token, brl, qty, price, trades in tokens_data:
    ws[f'A{row}'] = token
    ws[f'B{row}'] = brl
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'C{row}'] = qty
    ws[f'C{row}'].number_format = '0.00000000'
    ws[f'D{row}'] = price
    ws[f'D{row}'].number_format = '0.00000'
    ws[f'E{row}'] = trades
    row += 1

# Total row
ws[f'A{row}'] = "TOTAL"
ws[f'A{row}'].font = Font(bold=True)
ws[f'B{row}'] = 5815.50
ws[f'B{row}'].font = Font(bold=True)
ws[f'B{row}'].number_format = '#,##0.00'

# Ajustar larguras
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 15

# FOLHA 2: Historico OKX
ws2 = wb.create_sheet("OKX - Historico")
ws2['A1'] = "OKX TRADING HISTORY - CONVERSOES USDT-BRL"
ws2['A1'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
ws2['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws2.merge_cells('A1:E1')

ws2['A3'] = "Data"
ws2['B3'] = "Quantidade USDT"
ws2['C3'] = "Taxa BRL/USDT"
ws2['D3'] = "BRL Gasto"
ws2['E3'] = "Tipo"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws2[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws2[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

# Ler OKX
okx_file = "Mjk2NTE5NTE=~2023-05-09~2026-05-09~UTC+8~ea1e6f1509d9e505a9af7c3a52c37475_UnifiedBillHistory/OKX Trading History_2023-05-09~2026-05-09~UTC+8~ea1e6f1509d9e505a9af7c3a52c37475.csv"
row = 4
with open(okx_file, 'r', encoding='utf-8-sig') as f:
    next(f)
    reader = csv.DictReader(f)
    for r in reader:
        if r.get('Trade Type', '').strip() == 'Spot' and r.get('Symbol', '').strip() == 'USDT-BRL' and r.get('Action', '').strip() == 'Buy':
            amount = float(r.get('Amount', '0'))
            price = float(r.get('Filled Price', '0'))
            brl_spent = amount * price

            ws2[f'A{row}'] = r.get('Time', '')
            ws2[f'B{row}'] = amount
            ws2[f'B{row}'].number_format = '0.0000'
            ws2[f'C{row}'] = price
            ws2[f'C{row}'].number_format = '0.00000'
            ws2[f'D{row}'] = brl_spent
            ws2[f'D{row}'].number_format = '#,##0.00'
            ws2[f'E{row}'] = 'USDT-BRL'
            row += 1

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 15

# FOLHA 3: Historico Binance
ws3 = wb.create_sheet("Binance - Historico")
ws3['A1'] = "BINANCE FIAT PURCHASES - HISTORICO"
ws3['A1'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
ws3['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
ws3.merge_cells('A1:E1')

ws3['A3'] = "Data"
ws3['B3'] = "Token"
ws3['C3'] = "Quantidade"
ws3['D3'] = "BRL Gasto"
ws3['E3'] = "Taxa BRL/Unit"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws3[f'{col}3'].font = Font(bold=True, color='FFFFFF')
    ws3[f'{col}3'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

binance_file = "Binance-Histórico-de-Compras-com-Fiduciária-202605111343(UTC--3).csv"
row = 4
with open(binance_file, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()
    for i in range(1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        parts = line.split(',')
        if len(parts) < 8:
            continue

        valor_parts = parts[2].strip().split()
        valor_brl = float(valor_parts[0])

        receber_parts = parts[3].strip().split()
        qty = float(receber_parts[0])
        ticker = receber_parts[1] if len(receber_parts) > 1 else 'USDC'

        taxa = valor_brl / qty if qty > 0 else 0

        ws3[f'A{row}'] = parts[0]
        ws3[f'B{row}'] = ticker
        ws3[f'C{row}'] = qty
        ws3[f'C{row}'].number_format = '0.00000000'
        ws3[f'D{row}'] = valor_brl
        ws3[f'D{row}'].number_format = '#,##0.00'
        ws3[f'E{row}'] = taxa
        ws3[f'E{row}'].number_format = '0.00000'
        row += 1

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 18
ws3.column_dimensions['E'].width = 18

# Salvar
wb.save('Custo_Aquisicao_BRL_Lucas.xlsx')
print("COMPLETO: Custo_Aquisicao_BRL_Lucas.xlsx")
